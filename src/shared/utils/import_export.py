# -*- coding: utf-8 -*-
"""
数据导入导出模块
"""

import json
import csv
import pandas as pd
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
import io
from src.knowledge_management.domain.model.graph import KnowledgeGraph
from src.knowledge_management.domain.model.node import Node
from src.knowledge_management.domain.model.edge import Edge


class DataImportExport:
    """
    数据导入导出处理类
    """
    
    def __init__(self):
        """
        初始化导入导出处理器
        """
        self.supported_formats = ['json', 'csv', 'excel']
        
    def export_to_json(self, kg: KnowledgeGraph, filepath: Optional[str] = None, 
                      include_metadata: bool = True) -> Union[str, Dict[str, Any]]:
        """
        导出知识图谱到JSON格式
        
        Args:
            kg: 知识图谱实例
            filepath: 文件路径，如果为None则返回字典
            include_metadata: 是否包含元数据
            
        Returns:
            JSON字符串或字典
        """
        data = {
            'nodes': [node.to_dict() for node in kg.get_all_nodes()],
            'edges': [edge.to_dict() for edge in kg.get_all_edges()]
        }
        
        if include_metadata:
            stats = kg.get_statistics()
            data['metadata'] = {
                'export_timestamp': pd.Timestamp.now().isoformat(),
                'node_count': stats['node_count'],
                'edge_count': stats['edge_count'],
                'node_types': stats['node_types'],
                'edge_types': stats['edge_types'],
                'format_version': '1.0'
            }
            
        if filepath:
            try:
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                return filepath
            except Exception as e:
                raise ValueError(f"导出JSON文件失败: {str(e)}")
        else:
            return data
            
    def import_from_json(self, source: Union[str, Dict[str, Any], io.StringIO]) -> KnowledgeGraph:
        """
        从JSON格式导入知识图谱
        
        Args:
            source: JSON文件路径、字典或StringIO对象
            
        Returns:
            知识图谱实例
        """
        try:
            if isinstance(source, str):
                # 文件路径
                with open(source, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            elif isinstance(source, dict):
                # 字典
                data = source
            elif isinstance(source, io.StringIO):
                # StringIO对象
                data = json.load(source)
            else:
                raise ValueError("不支持的数据源类型")
                
            kg = KnowledgeGraph()
            
            # 导入节点
            if 'nodes' in data:
                for node_data in data['nodes']:
                    node = Node.from_dict(node_data)
                    kg.add_node(node)
                    
            # 导入边
            if 'edges' in data:
                for edge_data in data['edges']:
                    # 处理字段名兼容性
                    processed_edge_data = edge_data.copy()
                    
                    # 如果使用source/target字段，转换为source_id/target_id
                    if 'source' in processed_edge_data and 'source_id' not in processed_edge_data:
                        processed_edge_data['source_id'] = processed_edge_data['source']
                    if 'target' in processed_edge_data and 'target_id' not in processed_edge_data:
                        processed_edge_data['target_id'] = processed_edge_data['target']
                    
                    edge = Edge.from_dict(processed_edge_data)
                    kg.add_edge(edge)
                    
            return kg
            
        except Exception as e:
            raise ValueError(f"导入JSON数据失败: {str(e)}")
            
    def export_to_csv(self, kg: KnowledgeGraph, nodes_filepath: str, 
                     edges_filepath: str) -> tuple[str, str]:
        """
        导出知识图谱到CSV格式（分别导出节点和边）
        
        Args:
            kg: 知识图谱实例
            nodes_filepath: 节点CSV文件路径
            edges_filepath: 边CSV文件路径
            
        Returns:
            (节点文件路径, 边文件路径)
        """
        try:
            # 导出节点
            nodes_data = []
            for node in kg.get_all_nodes():
                node_dict = {
                    'id': node.id,
                    'label': node.label,
                    'type': node.type
                }
                
                # 添加位置信息
                if hasattr(node, 'x') and node.x is not None:
                    node_dict['x'] = node.x
                if hasattr(node, 'y') and node.y is not None:
                    node_dict['y'] = node.y
                    
                # 添加属性（扁平化）
                for key, value in node.properties.items():
                    if isinstance(value, (str, int, float, bool)):
                        node_dict[f'attr_{key}'] = value
                    else:
                        node_dict[f'attr_{key}'] = str(value)
                        
                nodes_data.append(node_dict)
                
            nodes_df = pd.DataFrame(nodes_data)
            nodes_df.to_csv(nodes_filepath, index=False, encoding='utf-8')
            
            # 导出边
            edges_data = []
            for edge in kg.get_all_edges():
                edge_dict = {
                        'source': edge.source_id,
                        'target': edge.target_id,
                        'type': edge.type
                    }
                
                # 添加属性（扁平化）
                for key, value in edge.properties.items():
                    if isinstance(value, (str, int, float, bool)):
                        edge_dict[f'attr_{key}'] = value
                    else:
                        edge_dict[f'attr_{key}'] = str(value)
                        
                edges_data.append(edge_dict)
                
            edges_df = pd.DataFrame(edges_data)
            edges_df.to_csv(edges_filepath, index=False, encoding='utf-8')
            
            return nodes_filepath, edges_filepath
            
        except Exception as e:
            raise ValueError(f"导出CSV文件失败: {str(e)}")
            
    def import_from_csv(self, nodes_filepath: str, edges_filepath: str) -> KnowledgeGraph:
        """
        从CSV格式导入知识图谱
        
        Args:
            nodes_filepath: 节点CSV文件路径
            edges_filepath: 边CSV文件路径
            
        Returns:
            知识图谱实例
        """
        try:
            kg = KnowledgeGraph()
            
            # 导入节点
            if Path(nodes_filepath).exists():
                nodes_df = pd.read_csv(nodes_filepath, encoding='utf-8')
                
                for _, row in nodes_df.iterrows():
                    # 基本信息
                    node_id = str(row['id'])
                    label = str(row['label'])
                    node_type = str(row['type'])
                    
                    # 位置信息
                    position = None
                    if 'x' in row and 'y' in row and pd.notna(row['x']) and pd.notna(row['y']):
                        position = (float(row['x']), float(row['y']))
                        
                    # 属性信息
                    attributes = {}
                    for col in row.index:
                        if col.startswith('attr_') and pd.notna(row[col]):
                            attr_name = col[5:]  # 移除 'attr_' 前缀
                            attributes[attr_name] = row[col]
                            
                    # 创建节点
                    node = Node(node_id, label, node_type, attributes)
                    if position:
                        node.set_position(position[0], position[1])
                        
                    kg.add_node(node)
                    
            # 导入边
            if Path(edges_filepath).exists():
                edges_df = pd.read_csv(edges_filepath, encoding='utf-8')
                
                for _, row in edges_df.iterrows():
                    # 基本信息
                    source = str(row['source'])
                    target = str(row['target'])
                    edge_type = str(row['type'])
                    
                    # 属性信息
                    attributes = {}
                    for col in row.index:
                        if col.startswith('attr_') and pd.notna(row[col]):
                            attr_name = col[5:]  # 移除 'attr_' 前缀
                            attributes[attr_name] = row[col]
                            
                    # 创建边
                    edge = Edge(source, target, edge_type, attributes)
                    kg.add_edge(edge)
                    
            return kg
            
        except Exception as e:
            raise ValueError(f"导入CSV数据失败: {str(e)}")
            
    def export_to_excel(self, kg: KnowledgeGraph, filepath: str) -> str:
        """
        导出知识图谱到Excel格式
        
        Args:
            kg: 知识图谱实例
            filepath: Excel文件路径
            
        Returns:
            文件路径
        """
        try:
            from openpyxl import Workbook
            
            # 创建工作簿
            wb = Workbook()
            
            # 创建节点工作表
            nodes_ws = wb.create_sheet(title='Nodes')
            
            # 导出节点
            nodes_data = []
            for node in kg.get_all_nodes():
                node_dict = {
                    'id': node.id,
                    'label': node.label,
                    'type': node.type
                }
                
                # 添加位置信息
                if hasattr(node, 'x') and node.x is not None:
                    node_dict['x'] = node.x
                if hasattr(node, 'y') and node.y is not None:
                    node_dict['y'] = node.y
                    
                # 添加属性（扁平化）
                for key, value in node.properties.items():
                    if isinstance(value, (str, int, float, bool)):
                        node_dict[f'attr_{key}'] = value
                    else:
                        node_dict[f'attr_{key}'] = str(value)
                        
                nodes_data.append(node_dict)
            
            # 写入节点数据
            if nodes_data:
                nodes_df = pd.DataFrame(nodes_data)
                for r_idx, row in enumerate(nodes_df.itertuples(index=False), 1):
                    for c_idx, value in enumerate(row, 1):
                        nodes_ws.cell(row=r_idx+1, column=c_idx, value=value)
                # 写入表头
                for c_idx, col_name in enumerate(nodes_df.columns, 1):
                    nodes_ws.cell(row=1, column=c_idx, value=col_name)
            else:
                # 空数据时写入表头
                nodes_ws.cell(row=1, column=1, value='id')
                nodes_ws.cell(row=1, column=2, value='label')
                nodes_ws.cell(row=1, column=3, value='type')
                
            # 创建边工作表
            edges_ws = wb.create_sheet(title='Edges')
            
            # 导出边
            edges_data = []
            for edge in kg.get_all_edges():
                edge_dict = {
                     'source': edge.source_id,
                     'target': edge.target_id,
                     'type': edge.type
                 }
                
                # 添加属性（扁平化）
                for key, value in edge.properties.items():
                    if isinstance(value, (str, int, float, bool)):
                        edge_dict[f'attr_{key}'] = value
                    else:
                        edge_dict[f'attr_{key}'] = str(value)
                        
                edges_data.append(edge_dict)
            
            # 写入边数据
            if edges_data:
                edges_df = pd.DataFrame(edges_data)
                for r_idx, row in enumerate(edges_df.itertuples(index=False), 1):
                    for c_idx, value in enumerate(row, 1):
                        edges_ws.cell(row=r_idx+1, column=c_idx, value=value)
                # 写入表头
                for c_idx, col_name in enumerate(edges_df.columns, 1):
                    edges_ws.cell(row=1, column=c_idx, value=col_name)
            else:
                # 空数据时写入表头
                edges_ws.cell(row=1, column=1, value='source')
                edges_ws.cell(row=1, column=2, value='target')
                edges_ws.cell(row=1, column=3, value='type')
                
            # 创建统计信息工作表
            stats_ws = wb.create_sheet(title='Statistics')
            
            # 导出统计信息
            stats = kg.get_statistics()
            stats_data = [
                {'Metric': 'Node Count', 'Value': stats['node_count']},
                {'Metric': 'Edge Count', 'Value': stats['edge_count']},
                {'Metric': 'Node Types', 'Value': ', '.join(stats['node_types'])},
                {'Metric': 'Edge Types', 'Value': ', '.join(stats['edge_types'])},
                {'Metric': 'Export Time', 'Value': pd.Timestamp.now().isoformat()}
            ]
            
            # 写入统计信息
            stats_ws.cell(row=1, column=1, value='Metric')
            stats_ws.cell(row=1, column=2, value='Value')
            for r_idx, item in enumerate(stats_data, 2):
                stats_ws.cell(row=r_idx, column=1, value=item['Metric'])
                stats_ws.cell(row=r_idx, column=2, value=item['Value'])
            
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 保存文件
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise ValueError(f"导出Excel文件失败: {str(e)}")
            
    def import_from_excel(self, filepath: str) -> KnowledgeGraph:
        """
        从Excel格式导入知识图谱
        
        Args:
            filepath: Excel文件路径
            
        Returns:
            知识图谱实例
        """
        try:
            kg = KnowledgeGraph()
            
            # 读取Excel文件
            excel_file = pd.ExcelFile(filepath)
            
            # 导入节点
            if 'Nodes' in excel_file.sheet_names:
                nodes_df = pd.read_excel(filepath, sheet_name='Nodes')
                
                for _, row in nodes_df.iterrows():
                    # 基本信息
                    node_id = str(row['id'])
                    label = str(row['label'])
                    node_type = str(row['type'])
                    
                    # 位置信息
                    position = None
                    if 'x' in row and 'y' in row and pd.notna(row['x']) and pd.notna(row['y']):
                        position = (float(row['x']), float(row['y']))
                        
                    # 属性信息
                    attributes = {}
                    for col in row.index:
                        if col.startswith('attr_') and pd.notna(row[col]):
                            attr_name = col[5:]  # 移除 'attr_' 前缀
                            attributes[attr_name] = row[col]
                            
                    # 创建节点
                    node = Node(node_id, label, node_type, attributes)
                    if position:
                        node.set_position(position[0], position[1])
                        
                    kg.add_node(node)
                    
            # 导入边
            if 'Edges' in excel_file.sheet_names:
                edges_df = pd.read_excel(filepath, sheet_name='Edges')
                
                for _, row in edges_df.iterrows():
                    # 基本信息
                    source = str(row['source'])
                    target = str(row['target'])
                    edge_type = str(row['type'])
                    
                    # 属性信息
                    attributes = {}
                    for col in row.index:
                        if col.startswith('attr_') and pd.notna(row[col]):
                            attr_name = col[5:]  # 移除 'attr_' 前缀
                            attributes[attr_name] = row[col]
                            
                    # 创建边
                    edge = Edge(source, target, edge_type, attributes)
                    kg.add_edge(edge)
                    
            return kg
            
        except Exception as e:
            raise ValueError(f"导入Excel数据失败: {str(e)}")
            
    def export_selection(self, kg: KnowledgeGraph, selected_nodes: List[str], 
                        format_type: str = 'json', filepath: Optional[str] = None) -> Union[str, Dict[str, Any]]:
        """
        导出选中的节点和相关边
        
        Args:
            kg: 知识图谱实例
            selected_nodes: 选中的节点ID列表
            format_type: 导出格式
            filepath: 文件路径
            
        Returns:
            文件路径或数据
        """
        # 创建包含选中节点的子图
        sub_kg = KnowledgeGraph()
        
        # 添加选中的节点
        for node_id in selected_nodes:
            node = kg.get_node(node_id)
            if node:
                sub_kg.add_node(node)
                
        # 添加相关的边
        for edge in kg.get_all_edges():
            if (edge.source in selected_nodes or edge.target in selected_nodes):
                # 确保两个节点都存在
                if kg.has_node(edge.source) and kg.has_node(edge.target):
                    if not sub_kg.has_node(edge.source):
                        sub_kg.add_node(kg.get_node(edge.source))
                    if not sub_kg.has_node(edge.target):
                        sub_kg.add_node(kg.get_node(edge.target))
                    sub_kg.add_edge(edge)
                    
        # 根据格式导出
        if format_type.lower() == 'json':
            return self.export_to_json(sub_kg, filepath)
        elif format_type.lower() == 'csv':
            if filepath:
                nodes_path = filepath.replace('.csv', '_nodes.csv')
                edges_path = filepath.replace('.csv', '_edges.csv')
                return self.export_to_csv(sub_kg, nodes_path, edges_path)
            else:
                raise ValueError("CSV导出需要指定文件路径")
        elif format_type.lower() == 'excel':
            if filepath:
                return self.export_to_excel(sub_kg, filepath)
            else:
                raise ValueError("Excel导出需要指定文件路径")
        else:
            raise ValueError(f"不支持的导出格式: {format_type}")
            
    def validate_import_data(self, data: Dict[str, Any]) -> Dict[str, List[str]]:
        """
        验证导入数据的有效性
        
        Args:
            data: 导入的数据字典
            
        Returns:
            验证结果字典，包含错误和警告信息
        """
        errors = []
        warnings = []
        
        # 检查基本结构
        if not isinstance(data, dict):
            errors.append("数据必须是字典格式")
            return {'errors': errors, 'warnings': warnings}
            
        if 'nodes' not in data and 'edges' not in data:
            errors.append("数据必须包含 'nodes' 或 'edges' 字段")
            
        # 验证节点数据
        if 'nodes' in data:
            if not isinstance(data['nodes'], list):
                errors.append("'nodes' 必须是列表格式")
            else:
                node_ids = set()
                for i, node in enumerate(data['nodes']):
                    if not isinstance(node, dict):
                        errors.append(f"节点 {i} 必须是字典格式")
                        continue
                        
                    # 检查必需字段
                    required_fields = ['id', 'label', 'type']
                    for field in required_fields:
                        if field not in node:
                            errors.append(f"节点 {i} 缺少必需字段: {field}")
                            
                    # 检查ID唯一性
                    if 'id' in node:
                        if node['id'] in node_ids:
                            errors.append(f"重复的节点ID: {node['id']}")
                        else:
                            node_ids.add(node['id'])
                            
        # 验证边数据
        if 'edges' in data:
            if not isinstance(data['edges'], list):
                errors.append("'edges' 必须是列表格式")
            else:
                for i, edge in enumerate(data['edges']):
                    if not isinstance(edge, dict):
                        errors.append(f"边 {i} 必须是字典格式")
                        continue
                        
                    # 检查必需字段
                    required_fields = ['source', 'target', 'type']
                    for field in required_fields:
                        if field not in edge:
                            errors.append(f"边 {i} 缺少必需字段: {field}")
                            
                    # 检查节点引用
                    if 'nodes' in data and 'source' in edge and 'target' in edge:
                        source_exists = any(n.get('id') == edge['source'] for n in data['nodes'])
                        target_exists = any(n.get('id') == edge['target'] for n in data['nodes'])
                        
                        if not source_exists:
                            warnings.append(f"边 {i} 引用了不存在的源节点: {edge['source']}")
                        if not target_exists:
                            warnings.append(f"边 {i} 引用了不存在的目标节点: {edge['target']}")
                            
        return {'errors': errors, 'warnings': warnings}
        
    def get_supported_formats(self) -> List[str]:
        """
        获取支持的文件格式
        
        Returns:
            支持的格式列表
        """
        return self.supported_formats.copy()
        
    def get_format_info(self, format_type: str) -> Dict[str, Any]:
        """
        获取格式信息
        
        Args:
            format_type: 格式类型
            
        Returns:
            格式信息字典
        """
        format_info = {
            'json': {
                'name': 'JSON',
                'description': '标准JSON格式，包含完整的节点和边信息',
                'extensions': ['.json'],
                'supports_metadata': True,
                'supports_attributes': True,
                'file_count': 1
            },
            'csv': {
                'name': 'CSV',
                'description': 'CSV格式，节点和边分别存储在不同文件中',
                'extensions': ['.csv'],
                'supports_metadata': False,
                'supports_attributes': True,
                'file_count': 2
            },
            'excel': {
                'name': 'Excel',
                'description': 'Excel格式，节点、边和统计信息存储在不同工作表中',
                'extensions': ['.xlsx', '.xls'],
                'supports_metadata': True,
                'supports_attributes': True,
                'file_count': 1
            }
        }
        
        return format_info.get(format_type.lower(), {})